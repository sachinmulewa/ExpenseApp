import streamlit as st
import openpyxl
from io import BytesIO

def create_excel_file(data):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet.cell(row=1, column=1, value="Home Expenses")
    for col_num, value in enumerate(data['home'], start=2):
        sheet.cell(row=1, column=col_num, value=value)

    sheet.cell(row=2, column=1, value="Travel Expenses")
    for col_num, value in enumerate(data['travel'], start=2):
        sheet.cell(row=2, column=col_num, value=value)

    sheet.cell(row=3, column=1, value="Food Expenses")
    for col_num, value in enumerate(data['food'], start=2):
        sheet.cell(row=3, column=col_num, value=value)

    # Save to a BytesIO stream
    excel_io = BytesIO()
    workbook.save(excel_io)
    excel_io.seek(0)
    return excel_io

# Streamlit UI
st.title("Expense Tracker")

with st.form("expense_form"):
    home_expenses = st.text_input("Enter Home Expenses (comma-separated)", "")
    travel_expenses = st.text_input("Enter Travel Expenses (comma-separated)", "")
    food_expenses = st.text_input("Enter Food Expenses (comma-separated)", "")
    file_name = st.text_input("Enter the file name:", "expenses.xlsx")
    submitted = st.form_submit_button("Generate Excel")

if submitted:
    data = {
        'home': [x.strip() for x in home_expenses.split(',') if x.strip()],
        'travel': [x.strip() for x in travel_expenses.split(',') if x.strip()],
        'food': [x.strip() for x in food_expenses.split(',') if x.strip()]
    }

    excel_file = create_excel_file(data)
    st.success("Excel file created successfully!")

    st.download_button(
        label="Download Excel File",
        data=excel_file,
        file_name=f"{file_name}.xlsx" if not file_name.endswith(".xlsx") else file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


